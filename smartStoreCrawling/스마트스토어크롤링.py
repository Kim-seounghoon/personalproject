import requests
#import urllib.request
import datetime # 날짜 형식 모듈
import sys
from openpyxl import Workbook # 엑셀을 사용하기 위한 모듈
from bs4 import BeautifulSoup
sys.setrecursionlimit(5000)
urls = [];
url = '';
while(url != '0'):
     url = input('url을 입력하세요(0은 종료) : ')
     urls.append(url)

wb = Workbook()    # Workbook을 생성하면 기본적으로 시트가 하나 생성됩니다.
ws = wb.active    # 시트 이름 바꾸기
ws.title = "sheet1"
column=1;
for i in range(0,len(urls)-1):    
    result = requests.get(urls[i])  #result.encoding = 'utf8'
    a = result.text # html 정보
    soup = BeautifulSoup(a, 'html.parser')    # 엑셀 Workbook 생성
    title = soup.select('head > title')
    name = soup.select('div > ul > li > a > div > div > strong')
    row=1
    cell = ws.cell(row, column)
    cell.value = (title[0].text).strip()
    row+=1
    for j in name:        
        cell = ws.cell(row, column)
        cell.value = j.text
        row+=1
        if(row == 6):
            break;
    column+=1
date=datetime.datetime.today().strftime("%m%d%H(%M_%S).xlsx")
wb.save(date);
print(date +'로 저장됨')
